# -*- coding: utf-8 -*-
"""
Kayıtlar Rotaları
"""

from datetime import datetime
from io import BytesIO
from flask import Blueprint, render_template, request, jsonify, send_file
from database import db, Kayit, Personel, Toplama, Order, AdetFiltresi, KayitAyrinti, kayit_ekle, kayit_guncelle, kayit_sil
from utils.excel_utils import build_template, load_excel_rows, calculate_file_hash
from utils.audit_utils import log_audit

kayitlar_bp = Blueprint('kayitlar', __name__, url_prefix='/kayitlar')


@kayitlar_bp.route('/')
def lista():
    """Kayıtlar listesi"""
    
    kayitlar = Kayit.query.order_by(Kayit.tarih.desc()).all()
    personeller = Personel.query.all()
    toplamalar = Toplama.query.all()
    
    return render_template('kayitlar.html',
                         kayitlar=kayitlar,
                         personeller=personeller,
                         toplamalar=toplamalar)


@kayitlar_bp.route('/api/list', methods=['GET'])
def api_list():
    """AJAX ile kayıtları getir"""
    from datetime import date as date_type

    # Filtreleri al
    personel_id = request.args.get('personel_id')
    toplama_id = request.args.get('toplama_id')
    tarih_baslangic_str = request.args.get('tarih_baslangic')
    tarih_bitis_str = request.args.get('tarih_bitis')

    # Tarih filtrelerini date objesine çevir (YYYY-MM-DD formatı beklenir)
    tarih_baslangic = None
    tarih_bitis = None
    try:
        if tarih_baslangic_str:
            tarih_baslangic = datetime.strptime(tarih_baslangic_str, '%Y-%m-%d').date()
        if tarih_bitis_str:
            tarih_bitis = datetime.strptime(tarih_bitis_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz tarih formatı!'}), 400

    # Sorgu oluştur (tarih dışındaki filtreler SQL ile)
    sorgu = Kayit.query

    if personel_id:
        sorgu = sorgu.filter_by(personel_id=int(personel_id))

    if toplama_id:
        sorgu = sorgu.filter_by(toplama_id=int(toplama_id))

    kayitlar = sorgu.all()

    # Tarihler DD.MM.YYYY string olarak saklandığından Python seviyesinde filtrele
    if tarih_baslangic or tarih_bitis:
        filtrelenmis = []
        for k in kayitlar:
            try:
                k_tarih = datetime.strptime(k.tarih, '%d.%m.%Y').date()
            except ValueError:
                continue
            if tarih_baslangic and k_tarih < tarih_baslangic:
                continue
            if tarih_bitis and k_tarih > tarih_bitis:
                continue
            filtrelenmis.append(k)
        kayitlar = filtrelenmis

    # Tarihe göre azalan sıralama
    kayitlar.sort(key=lambda k: (
        datetime.strptime(k.tarih, '%d.%m.%Y') if '.' in k.tarih else datetime.min
    ), reverse=True)

    return jsonify({
        'basarili': True,
        'kayitlar': [k.to_dict() for k in kayitlar],
        'toplam': len(kayitlar)
    })


@kayitlar_bp.route('/api/guncelle/<int:id>', methods=['PUT'])
def api_guncelle(id):
    """AJAX ile kayıt güncelle"""
    try:
        veri = request.json or {}
        kayit = Kayit.query.get(id)
        if not kayit:
            return jsonify({'basarili': False, 'mesaj': 'Kayıt bulunamadı!'}), 404

        eski = kayit.to_dict()

        tarih_raw = veri.get('tarih', kayit.tarih)
        if '-' in str(tarih_raw) and len(str(tarih_raw)) == 10:
            parts = str(tarih_raw).split('-')
            tarih_formatted = f"{parts[2]}.{parts[1]}.{parts[0]}"
        else:
            tarih_formatted = tarih_raw

        sonuc = kayit_guncelle(
            id=id,
            tarih=tarih_formatted,
            personel_id=int(veri['personel_id']) if veri.get('personel_id') else None,
            toplama_id=int(veri.get('toplama_id', kayit.toplama_id)),
            trendyol_siparis=float(veri.get('trendyol_siparis', kayit.trendyol_siparis) or 0),
            trendyol_fatura=float(veri.get('trendyol_fatura', kayit.trendyol_fatura) or 0),
            diger_pazar=float(veri.get('diger_pazar', kayit.diger_pazar) or 0),
            not_alan=veri.get('not', kayit.not_alan or ''),
        )

        log_audit('update', 'kayitlar', id, eski_deger=eski, yeni_deger=Kayit.query.get(id).to_dict())
        return jsonify(sonuc)

    except Exception:
        return jsonify({'basarili': False, 'mesaj': 'İşlem sırasında bir hata oluştu.'}), 500


@kayitlar_bp.route('/api/sil/<int:id>', methods=['DELETE'])
def api_sil(id):
    """AJAX ile kayıt sil"""
    try:
        kayit = Kayit.query.get(id)
        if kayit:
            log_audit('delete', 'kayitlar', id, eski_deger=kayit.to_dict())
        sonuc = kayit_sil(id)
        return jsonify(sonuc)
    except Exception as e:
        return jsonify({'basarili': False, 'mesaj': f'Hata: {str(e)}'}), 500


@kayitlar_bp.route('/api/template')
def download_template():
    """Günlük kayıt Excel şablonu indir"""
    headers = ['Tarih', 'Personel', 'Toplama', 'Trendyol Siparis', 'Trendyol Fatura', 'Diger Pazar', 'Not']
    stream = build_template(headers)
    return send_file(
        stream,
        as_attachment=True,
        download_name='gunluk_kayit_sablonu.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@kayitlar_bp.route('/api/excel-yukle', methods=['POST'])
def excel_yukle():
    """Excel ile toplu kayıt yükleme"""
    file = request.files.get('file')
    if not file:
        return jsonify({'basarili': False, 'mesaj': 'Dosya zorunludur!'}), 400

    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz dosya formatı! (.xlsx gerekli)'}), 400

    expected_headers = ['Tarih', 'Personel', 'Toplama', 'Trendyol Siparis', 'Trendyol Fatura', 'Diger Pazar', 'Not']

    try:
        headers, rows = load_excel_rows(file)
    except Exception:
        return jsonify({'basarili': False, 'mesaj': 'Excel okunamadı. Dosya formatını kontrol edin.'}), 400

    if headers != expected_headers:
        return jsonify({
            'basarili': False,
            'mesaj': 'Excel başlıkları hatalı!',
            'beklenen': expected_headers,
            'bulunan': headers,
        }), 400

    basarili = 0
    hatalar = []

    for idx, row in enumerate(rows, start=2):
        try:
            tarih_raw = str(row.get('Tarih', '')).strip()
            personel_adi = str(row.get('Personel', '')).strip()
            toplama_adi = str(row.get('Toplama', '')).strip()

            if not tarih_raw or not toplama_adi:
                hatalar.append({'satir': idx, 'hata': 'Tarih veya Toplama boş'})
                continue

            personel_id = None
            if personel_adi:
                personel = Personel.query.filter_by(ad=personel_adi).first()
                if not personel:
                    hatalar.append({'satir': idx, 'hata': f'Personel bulunamadı: {personel_adi}'})
                    continue
                personel_id = personel.id

            toplama = Toplama.query.filter_by(ad=toplama_adi).first()
            if not toplama:
                hatalar.append({'satir': idx, 'hata': f'Toplama bulunamadı: {toplama_adi}'})
                continue

            # Format tarih
            if '-' in tarih_raw and len(tarih_raw) == 10:
                parts = tarih_raw.split('-')
                tarih_formatted = f"{parts[2]}.{parts[1]}.{parts[0]}"
            else:
                tarih_formatted = tarih_raw

            from database import kayit_ekle as _kayit_ekle
            _kayit_ekle(
                tarih=tarih_formatted,
                personel_id=personel_id,
                toplama_id=toplama.id,
                trendyol_siparis=float(row.get('Trendyol Siparis') or 0),
                trendyol_fatura=float(row.get('Trendyol Fatura') or 0),
                diger_pazar=float(row.get('Diger Pazar') or 0),
                not_alan=str(row.get('Not', '') or ''),
            )
            basarili += 1
        except Exception:
            hatalar.append({'satir': idx, 'hata': 'Satır işlenemedi'})

    log_audit('excel_toplu_yukleme', 'kayitlar', None, yeni_deger={'basarili': basarili, 'hatali': len(hatalar)})

    return jsonify({
        'basarili': True,
        'mesaj': f'{basarili} kayıt eklendi, {len(hatalar)} satır atlandı.',
        'ozet': {
            'toplam_satir': len(rows),
            'basarili': basarili,
            'hatali': len(hatalar),
        },
        'hatalar': hatalar,
    })


@kayitlar_bp.route('/api/senkronize', methods=['POST'])
def senkronize_et():
    """Seçili tarihin siparişlerini kayıtlara senkronize et.

    Sadece seçili tarihteki BEKLEMEDE siparişleri işler.
    Sipariş No bazında gruplama yapar.  Her grubun ilk satırında (grup_baslangic_satiri=True)
    olan siparişin personel_id'si grubun personeli olarak kullanılır.
    AdetFiltresi varsa aynı (urun_kodu, beden) için farklı adet aralıklarına ayrı
    KayitAyrinti satırları oluşturulur.
    Sadece başarıyla kayda alınan siparişleri TAMAMLANDI olarak işaretler.
    """
    try:
        veri = request.get_json() or {}
        tarih_str = str(veri.get('tarih') or '').strip()
        if not tarih_str:
            return jsonify({'basarili': False, 'mesaj': 'Lütfen senkronize edilecek tarihi seçin!'}), 400

        try:
            tarih_obj = datetime.strptime(tarih_str, '%Y-%m-%d').date()
        except ValueError:
            return jsonify({'basarili': False, 'mesaj': 'Geçersiz tarih formatı!'}), 400

        tarih_db = tarih_obj.strftime('%d.%m.%Y')
        gun_baslangici = datetime.combine(tarih_obj, datetime.min.time())
        gun_bitis = datetime.combine(tarih_obj, datetime.max.time())

        siparisler = Order.query.filter(
            Order.durum == 'BEKLEMEDE',
            Order.tarih >= gun_baslangici,
            Order.tarih <= gun_bitis,
        ).all()

        if not siparisler:
            return jsonify({
                'basarili': False,
                'mesaj': f'{tarih_db} tarihi için BEKLEMEDE sipariş bulunamadı.',
                'yeni': 0,
                'guncellendi': 0,
                'eklenen_adet': 0,
            })

        now = datetime.now()
        sonuc = _siparisleri_kayitlara_aktar(
            siparisler=siparisler,
            tarih_db=tarih_db,
            now=now,
        )

        db.session.commit()
        log_audit('senkronizasyon', 'kayitlar', None,
                  yeni_deger={
                      'tarih': tarih_db,
                      'islenen_siparis': sonuc['islenen_siparis'],
                      'eklenen_adet': sonuc['eklenen_adet'],
                      'yeni': sonuc['yeni'],
                      'guncellendi': sonuc['guncellendi'],
                  })

        return jsonify({
            'basarili': True,
            'mesaj': (
                f'{sonuc["islenen_siparis"]} sipariş işlendi, '
                f'{sonuc["eklenen_adet"]} adet senkronize edildi. '
                f'{sonuc["yeni"]} yeni kayıt oluşturuldu, {sonuc["guncellendi"]} kayıt güncellendi.'
            ),
            'tarih': tarih_db,
            'eklenen_adet': sonuc['eklenen_adet'],
            'yeni': sonuc['yeni'],
            'guncellendi': sonuc['guncellendi'],
        })

    except Exception:
        db.session.rollback()
        import traceback; traceback.print_exc()
        return jsonify({'basarili': False, 'mesaj': 'Senkronizasyon sırasında bir hata oluştu.'}), 500

@kayitlar_bp.route('/api/excel-export', methods=['POST'])
def excel_export():
    """Kayıtları Excel'e aktar (Tarih aralığına göre, ürün detaylarıyla)"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    veri = request.get_json() or {}
    baslangic_str = str(veri.get('baslangic_tarihi') or '').strip()
    bitis_str = str(veri.get('bitis_tarihi') or '').strip()

    if not baslangic_str or not bitis_str:
        return jsonify({'basarili': False, 'mesaj': 'Tarih aralığı zorunludur!'}), 400

    try:
        tarih_bas = datetime.strptime(baslangic_str, '%Y-%m-%d').date()
        tarih_bit = datetime.strptime(bitis_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz tarih formatı!'}), 400

    tum_kayitlar = Kayit.query.all()
    kayitlar = []
    for k in tum_kayitlar:
        try:
            k_tarih = datetime.strptime(k.tarih, '%d.%m.%Y').date()
        except ValueError:
            continue
        if tarih_bas <= k_tarih <= tarih_bit:
            kayitlar.append(k)

    kayitlar.sort(key=lambda k: datetime.strptime(k.tarih, '%d.%m.%Y'), reverse=True)

    if not kayitlar:
        return jsonify({'basarili': False, 'mesaj': 'Seçili tarih aralığında kayıt bulunamadı!'}), 404

    wb = Workbook()
    ws = wb.active
    ws.title = 'Kayıtlar'

    basliklar = ['Tarih', 'Personel', 'Toplama', 'Sipariş Sayısı', 'Ürün Kodu', 'Beden', 'Adet', 'Senkronizasyon Tarihi']
    ws.append(basliklar)

    header_fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for kayit in kayitlar:
        personel_adi = kayit.personel.ad if kayit.personel else '—'
        toplama_adi = kayit.toplama.ad if kayit.toplama else '—'
        siparis_sayisi = int(kayit.trendyol_siparis or 0)
        senk_tarihi = kayit.son_senkronizasyon.strftime('%d.%m.%Y %H:%M') if kayit.son_senkronizasyon else '—'

        ayrintilari = KayitAyrinti.query.filter_by(kayit_id=kayit.id).all()

        if not ayrintilari:
            ws.append([kayit.tarih, personel_adi, toplama_adi, siparis_sayisi, '—', '—', '—', senk_tarihi])
        else:
            for i, ayrinti in enumerate(ayrintilari):
                if i == 0:
                    ws.append([
                        kayit.tarih, personel_adi, toplama_adi, siparis_sayisi,
                        ayrinti.urun_kodu, ayrinti.beden or '—', ayrinti.adet, senk_tarihi,
                    ])
                else:
                    ws.append(['', '', '', '', ayrinti.urun_kodu, ayrinti.beden or '—', ayrinti.adet, ''])

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 22

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    dosya_adi = f'kayitlar_{datetime.now().strftime("%d_%m_%Y")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=dosya_adi,
    )


def _senkronize_upload(upload_id):
    """Belirli bir upload_id için senkronizasyon yap (route dışından çağrılabilir)"""
    now = datetime.now()
    tarih_db = now.strftime('%d.%m.%Y')

    siparisler = Order.query.filter_by(
        excel_yukleme_id=upload_id,
        senkronize_edildi=False,
    ).filter(Order.durum == 'BEKLEMEDE').all()

    if not siparisler:
        return {'basarili': True, 'mesaj': 'Senkronize edilecek BEKLEMEDE sipariş bulunamadı.', 'yeni': 0, 'guncellendi': 0}

    sonuc = _siparisleri_kayitlara_aktar(
        siparisler=siparisler,
        tarih_db=tarih_db,
        now=now,
    )

    db.session.commit()
    return {
        'basarili': True,
        'mesaj': f'{sonuc["yeni"]} yeni kayıt oluşturuldu, {sonuc["guncellendi"]} kayıt güncellendi.',
        'yeni': sonuc['yeni'],
        'guncellendi': sonuc['guncellendi'],
    }


def _adet_filtre_araligi_bul(toplama_id, urun_kodu, beden, siparis_adedi):
    filtreler = AdetFiltresi.query.filter_by(
        toplama_id=toplama_id,
        urun_kodu=urun_kodu,
        beden=beden or None,
    ).order_by(AdetFiltresi.min_adet).all()

    if not filtreler and beden:
        filtreler = AdetFiltresi.query.filter_by(
            toplama_id=toplama_id,
            urun_kodu=urun_kodu,
            beden=None,
        ).order_by(AdetFiltresi.min_adet).all()

    min_f = None
    max_f = None
    for filtre in filtreler:
        ust = filtre.max_adet if filtre.max_adet is not None else float('inf')
        if filtre.min_adet <= siparis_adedi <= ust:
            min_f = filtre.min_adet
            max_f = filtre.max_adet
            break

    return min_f, max_f


def _siparisleri_kayitlara_aktar(siparisler, tarih_db, now):
    toplama_gruplari = {}

    for siparis in siparisler:
        if siparis.toplama_id is None:
            continue

        toplama_grubu = toplama_gruplari.setdefault(siparis.toplama_id, {
            'siparisler': {},
        })

        siparis_grubu = toplama_grubu['siparisler'].setdefault(siparis.siparis_no, {
            'satirlar': [],
            'personel_id': None,
        })
        siparis_grubu['satirlar'].append(siparis)

        if siparis.grup_baslangic_satiri and siparis.personel_id is not None:
            siparis_grubu['personel_id'] = siparis.personel_id
        elif siparis_grubu['personel_id'] is None and siparis.personel_id is not None:
            siparis_grubu['personel_id'] = siparis.personel_id

    yeni = 0
    guncellendi = 0
    eklenen_adet = 0
    islenen_siparis = 0

    for toplama_id, toplama_grubu in toplama_gruplari.items():
        siparis_gruplari = toplama_grubu['siparisler']
        if not siparis_gruplari:
            continue

        islenen_siparis += len(siparis_gruplari)
        ilk_siparis = next(iter(siparis_gruplari.values()))
        grup_personel_id = ilk_siparis['personel_id']

        kayit = Kayit.query.filter_by(
            tarih=tarih_db,
            toplama_id=toplama_id,
            personel_id=grup_personel_id,
        ).first()

        if kayit:
            kayit.senkronizasyon_sayisi = (kayit.senkronizasyon_sayisi or 0) + 1
            kayit.son_senkronizasyon = now
            guncellendi += 1
        else:
            kayit = Kayit(
                tarih=tarih_db,
                personel_id=grup_personel_id,
                toplama_id=toplama_id,
                trendyol_siparis=0,
                trendyol_fatura=0,
                diger_pazar=0,
                not_alan='',
                senkronizasyon_sayisi=1,
                son_senkronizasyon=now,
            )
            db.session.add(kayit)
            db.session.flush()
            yeni += 1

        for siparis_grubu in siparis_gruplari.values():
            for siparis in siparis_grubu['satirlar']:
                siparis_adedi = int(siparis.adet or 0)
                urun_kodu = (siparis.urun.ana_kod if siparis.urun else siparis.urun_kodu_ham) or ''
                beden = siparis.beden or ''
                min_f, max_f = _adet_filtre_araligi_bul(
                    toplama_id=toplama_id,
                    urun_kodu=urun_kodu,
                    beden=beden,
                    siparis_adedi=siparis_adedi,
                )

                db.session.add(KayitAyrinti(
                    kayit_id=kayit.id,
                    urun_kodu=urun_kodu,
                    beden=beden or None,
                    adet=siparis_adedi,
                    min_adet_filtre=min_f,
                    max_adet_filtre=max_f,
                    olusturulma_tarihi=now,
                ))

                siparis.durum = 'TAMAMLANDI'
                siparis.senkronize_edildi = True
                siparis.senkronize_tarihi = now
                eklenen_adet += siparis_adedi

        kayit.trendyol_siparis = (kayit.trendyol_siparis or 0) + len(siparis_gruplari)

    return {
        'yeni': yeni,
        'guncellendi': guncellendi,
        'eklenen_adet': eklenen_adet,
        'islenen_siparis': islenen_siparis,
    }
