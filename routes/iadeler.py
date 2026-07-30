# -*- coding: utf-8 -*-
"""İade yönetimi rotaları"""

from io import BytesIO
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, jsonify, send_file
from openpyxl import Workbook
from sqlalchemy import or_

from database import db, Return, Product, Toplama, ExcelUpload
from routes.excel_handler import (
    validate_excel_upload,
    match_product_by_code,
    determine_toplama,
    parse_row_date,
)
from utils.excel_utils import build_template, load_excel_rows, calculate_file_hash
from utils.audit_utils import log_audit

iadeler_bp = Blueprint('iadeler', __name__, url_prefix='/iadeler')

IADE_HEADERS = ['Siparis No', 'Tarih', 'Urun Kodu', 'Beden', 'Adet', 'Sebep']


# ── Yardımcı Fonksiyonlar ─────────────────────────────────────────────────────

def _check_return(siparis_no, tarih, urun_id, urun_kodu_ham, beden, adet, toplama_id, product=None):
    """Returns list of error strings for a return record."""
    errors = []
    if not siparis_no or str(siparis_no).strip() == '':
        errors.append('Sipariş No boş')
    if not tarih:
        errors.append('Tarih boş')
    if not urun_id:
        if urun_kodu_ham:
            errors.append(f'Ürün tanımsız: {urun_kodu_ham}')
        else:
            errors.append('Ürün Kodu boş')
    if product and product.beden_ayrimi and not beden:
        errors.append('Beden boş/tanımsız')
    if not toplama_id:
        errors.append('Toplama seçilmemiş')
    if not adet or adet <= 0:
        errors.append('Adet ≤ 0')
    return errors


def _revalidate_return(ret):
    """HATALI iade kaydını yeniden doğrula ve ürün/toplama eşlemesini güncelle."""
    raw_code = str(ret.urun_kodu_ham or '').strip()
    product = None

    if raw_code:
        product = match_product_by_code(raw_code)
        ret.urun_id = product.id if product else None
    elif ret.urun_id:
        product = Product.query.get(ret.urun_id)

    beden_error = None
    if product:
        toplama_id, beden_error = determine_toplama(product, ret.beden)
        if toplama_id:
            ret.toplama_id = toplama_id

    errors = _check_return(
        siparis_no=ret.siparis_no,
        tarih=ret.tarih,
        urun_id=ret.urun_id,
        urun_kodu_ham=ret.urun_kodu_ham,
        beden=ret.beden,
        adet=ret.adet,
        toplama_id=ret.toplama_id,
        product=product,
    )
    if beden_error and beden_error not in errors:
        errors.append(beden_error)
    return errors


# ── Routes ────────────────────────────────────────────────────────────────────

@iadeler_bp.route('/')
def index():
    return render_template('iadeler.html')


@iadeler_bp.route('/api/meta')
def api_meta():
    """Dropdown verileri: toplamalar"""
    toplamalar = Toplama.query.order_by(Toplama.ad).all()
    return jsonify({
        'basarili': True,
        'toplamalar': [{'id': t.id, 'ad': t.ad} for t in toplamalar],
    })


@iadeler_bp.route('/api/template')
def download_template():
    stream = build_template(IADE_HEADERS)
    return send_file(stream, as_attachment=True, download_name='iade_template.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@iadeler_bp.route('/api/excel-yukle', methods=['POST'])
def excel_yukle():
    file = request.files.get('file')
    if not file:
        return jsonify({'basarili': False, 'mesaj': 'Dosya zorunludur!'}), 400

    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz dosya formatı! (.xlsx gerekli)'}), 400

    file_hash = calculate_file_hash(file)
    if ExcelUpload.query.filter_by(dosya_hash=file_hash).first():
        return jsonify({'basarili': False, 'mesaj': 'Bu Excel dosyası daha önce yüklenmiş!'}), 409

    try:
        headers, rows = load_excel_rows(file)
    except Exception:
        return jsonify({'basarili': False, 'mesaj': 'Excel okunamadı. Dosya formatını kontrol edin.'}), 400

    actual_core = [str(h).strip() for h in headers[:len(IADE_HEADERS)]]
    if actual_core != IADE_HEADERS:
        return jsonify({
            'basarili': False,
            'mesaj': 'Excel başlıkları/sırası hatalı!',
            'beklenen': IADE_HEADERS,
            'bulunan': headers,
        }), 400

    upload = ExcelUpload(
        modul='iade', dosya_adi=filename,
        dosya_hash=file_hash, toplam_satir=len(rows),
    )
    db.session.add(upload)
    db.session.flush()
    log_audit('excel_yukleme_basladi', 'excel_uploads', upload.id,
              yeni_deger={'modul': 'iade', 'dosya': filename})

    basarili = 0
    hatali = 0
    row_errors = []
    unique_returns = set()
    seen_siparis_nos = set()

    for row_no, row in enumerate(rows, start=2):
        siparis_no = str(row.get('Siparis No', '') or '').strip()
        urun_kodu_raw = str(row.get('Urun Kodu', '') or '').strip()
        beden = str(row.get('Beden', '') or '').strip() or None
        sebebi = str(row.get('Sebep', '') or '').strip() or None

        try:
            adet = int(float(row.get('Adet') or 0))
        except (ValueError, TypeError):
            adet = 0

        tarih = parse_row_date(row.get('Tarih'))

        errors = []
        product = None
        urun_id = None
        toplama_id = None

        if not siparis_no:
            errors.append('Sipariş No boş')

        if not urun_kodu_raw:
            errors.append('Ürün Kodu boş')
        else:
            product = match_product_by_code(urun_kodu_raw)
            if not product:
                errors.append(f'Ürün tanımsız: {urun_kodu_raw}')
            else:
                urun_id = product.id
                tid, beden_error = determine_toplama(product, beden)
                if beden_error:
                    errors.append(beden_error)
                else:
                    toplama_id = tid

        if not toplama_id and not any('Toplama' in e or 'beden' in e.lower() or 'Beden' in e for e in errors):
            errors.append('Toplama seçilmemiş')

        if adet <= 0:
            errors.append('Adet ≤ 0')

        if not tarih:
            errors.append('Tarih boş')

        if errors:
            durum = 'HATALI'
            hata_sebebi = '; '.join(errors)
            hatali += 1
            row_errors.append({
                'satir': row_no,
                'siparis_no': siparis_no,
                'urun_kodu': urun_kodu_raw,
                'beden': beden or '',
                'adet': adet,
                'hata': hata_sebebi,
            })
        else:
            durum = 'BEKLEMEDE'
            hata_sebebi = None
            basarili += 1

        # Aynı sipariş no + urun kodu ile HATALI kayıt varsa sil (yeniden yükleme)
        if siparis_no:
            existing_q = Return.query.filter(
                Return.siparis_no == siparis_no,
                Return.durum == 'HATALI',
            )
            if urun_kodu_raw:
                existing_q = existing_q.filter(
                    or_(Return.urun_kodu_ham == urun_kodu_raw,
                        Return.urun_id == urun_id) if urun_id else Return.urun_kodu_ham == urun_kodu_raw
                )
            existing = existing_q.first()
            if existing:
                db.session.delete(existing)
                db.session.flush()

        ret = Return(
            siparis_no=siparis_no or '?',
            tarih=tarih or datetime.utcnow(),
            urun_id=urun_id,
            urun_kodu_ham=urun_kodu_raw or None,
            beden=beden,
            adet=adet,
            sebebi=sebebi,
            toplama_id=toplama_id,
            durum=durum,
            hata_sebebi=hata_sebebi,
            excel_yukleme_id=upload.id,
        )
        db.session.add(ret)
        if siparis_no:
            seen_siparis_nos.add(siparis_no)
            unique_returns.add(siparis_no)

    upload.basarili = basarili
    upload.basarisiz = hatali
    log_audit('excel_yukleme_tamamlandi', 'returns', upload.id, yeni_deger={
        'toplam_satir': len(rows),
        'basarili': basarili,
        'hatali': hatali,
    })
    db.session.commit()

    return jsonify({
        'basarili': True,
        'mesaj': 'İade yükleme tamamlandı',
        'upload_id': upload.id,
        'ozet': {
            'toplam_satir': len(rows),
            'tekil_siparis': len(unique_returns),
            'islenen_satir': basarili,
            'hatali_satir': hatali,
        },
        'hatalar': row_errors,
    })


@iadeler_bp.route('/api/list')
def api_list():
    page = max(int(request.args.get('page', 1)), 1)
    per_page = min(max(int(request.args.get('per_page', 20)), 1), 100)

    siparis_no = request.args.get('siparis_no', '').strip()
    urun_kodu = request.args.get('urun_kodu', '').strip()
    tarih_baslangic = request.args.get('tarih_baslangic', '').strip()
    tarih_bitis = request.args.get('tarih_bitis', '').strip()
    f_toplama_id = request.args.get('toplama_id', '').strip()
    f_durum = request.args.get('durum', '').strip()

    query = Return.query

    if siparis_no:
        query = query.filter(Return.siparis_no.ilike(f'%{siparis_no}%'))

    if urun_kodu:
        query = query.outerjoin(Product, Return.urun_id == Product.id).filter(
            or_(
                Product.ana_kod.ilike(f'%{urun_kodu}%'),
                Return.urun_kodu_ham.ilike(f'%{urun_kodu}%'),
            )
        )

    if tarih_baslangic:
        try:
            query = query.filter(Return.tarih >= datetime.strptime(tarih_baslangic, '%Y-%m-%d'))
        except ValueError:
            pass

    if tarih_bitis:
        try:
            dt_end = datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Return.tarih < dt_end)
        except ValueError:
            pass

    if f_toplama_id:
        query = query.filter(Return.toplama_id == int(f_toplama_id))

    if f_durum:
        query = query.filter(Return.durum == f_durum)

    pagination = query.order_by(Return.id.desc()).paginate(page=page, per_page=per_page, error_out=False)

    rows = []
    for item in pagination.items:
        rows.append({
            'id': item.id,
            'siparis_no': item.siparis_no,
            'tarih': item.tarih.strftime('%Y-%m-%d') if item.tarih else None,
            'urun_kodu': (item.urun.ana_kod if item.urun else item.urun_kodu_ham) or '',
            'beden': item.beden or '',
            'adet': item.adet,
            'sebep': item.sebebi or '',
            'toplama': item.toplama.ad if item.toplama else '',
            'toplama_id': item.toplama_id,
            'durum': item.durum or 'BEKLEMEDE',
            'hata_sebebi': item.hata_sebebi or '',
        })

    return jsonify({'basarili': True, 'kayitlar': rows, 'toplam': pagination.total})


@iadeler_bp.route('/api/<int:return_id>', methods=['GET'])
def api_get(return_id):
    ret = Return.query.get_or_404(return_id)
    return jsonify({
        'basarili': True,
        'iade': {
            'id': ret.id,
            'siparis_no': ret.siparis_no,
            'tarih': ret.tarih.strftime('%Y-%m-%d') if ret.tarih else None,
            'urun_kodu': (ret.urun.ana_kod if ret.urun else ret.urun_kodu_ham) or '',
            'beden': ret.beden or '',
            'adet': ret.adet,
            'sebebi': ret.sebebi or '',
            'toplama_id': ret.toplama_id,
            'durum': ret.durum or 'BEKLEMEDE',
            'hata_sebebi': ret.hata_sebebi or '',
        },
    })


@iadeler_bp.route('/api/<int:return_id>', methods=['PUT'])
def api_guncelle(return_id):
    ret = Return.query.get_or_404(return_id)
    data = request.json or {}
    eski_durum = ret.durum
    eski = {'siparis_no': ret.siparis_no, 'beden': ret.beden, 'adet': ret.adet, 'sebebi': ret.sebebi, 'durum': ret.durum}

    if 'beden' in data:
        ret.beden = str(data['beden'] or '').strip() or None
    if 'adet' in data:
        adet = int(float(data['adet'] or 0))
        if adet <= 0:
            return jsonify({'basarili': False, 'mesaj': 'Adet 0 veya negatif olamaz'}), 400
        ret.adet = adet
    if 'sebebi' in data:
        ret.sebebi = str(data['sebebi'] or '').strip() or None
    if 'toplama_id' in data and data['toplama_id']:
        ret.toplama_id = int(data['toplama_id'])
    if 'tarih' in data and data['tarih']:
        try:
            ret.tarih = datetime.strptime(str(data['tarih']), '%Y-%m-%d')
        except ValueError:
            pass
    if 'urun_kodu' in data and data['urun_kodu']:
        urun_kodu = str(data['urun_kodu']).strip()
        ret.urun_kodu_ham = urun_kodu
        product = match_product_by_code(urun_kodu)
        if product:
            ret.urun_id = product.id
            if 'toplama_id' not in data or not data['toplama_id']:
                tid, _ = determine_toplama(product, ret.beden)
                if tid:
                    ret.toplama_id = tid
        else:
            ret.urun_id = None

    # Durum otomatik belirleme (TAMAMLANDI ise değiştirme)
    if eski_durum != 'TAMAMLANDI':
        errors = _check_return(
            siparis_no=ret.siparis_no,
            tarih=ret.tarih,
            urun_id=ret.urun_id,
            urun_kodu_ham=ret.urun_kodu_ham,
            beden=ret.beden,
            adet=ret.adet,
            toplama_id=ret.toplama_id,
            product=ret.urun,
        )
        if errors:
            ret.durum = 'HATALI'
            ret.hata_sebebi = '; '.join(errors)
        else:
            ret.durum = 'BEKLEMEDE'
            ret.hata_sebebi = None

    log_audit('update', 'returns', return_id, eski_deger=eski,
              yeni_deger={'beden': ret.beden, 'adet': ret.adet, 'sebebi': ret.sebebi, 'durum': ret.durum})
    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': 'İade güncellendi', 'yeni_durum': ret.durum})


@iadeler_bp.route('/api/<int:return_id>', methods=['DELETE'])
def api_sil(return_id):
    ret = Return.query.get_or_404(return_id)
    log_audit('delete', 'returns', return_id, eski_deger={'siparis_no': ret.siparis_no})
    db.session.delete(ret)
    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': 'İade silindi'})


@iadeler_bp.route('/api/grup-sil', methods=['POST'])
def api_grup_sil():
    """Aynı sipariş no'suna sahip tüm iadeleri sil"""
    data = request.json or {}
    siparis_no = str(data.get('siparis_no', '')).strip()
    if not siparis_no:
        return jsonify({'basarili': False, 'mesaj': 'Sipariş No gerekli'})
    silinen = Return.query.filter(Return.siparis_no == siparis_no).delete(synchronize_session=False)
    db.session.commit()
    log_audit('grup_sil', 'returns', None, yeni_deger={'siparis_no': siparis_no, 'silinen': silinen})
    return jsonify({'basarili': True, 'mesaj': f'{silinen} iade silindi'})


@iadeler_bp.route('/api/grup-revalidate', methods=['POST'])
def api_grup_revalidate():
    """Aynı sipariş no'suna sahip HATALI iadeleri yeniden doğrula"""
    data = request.json or {}
    siparis_no = str(data.get('siparis_no', '')).strip()
    if not siparis_no:
        return jsonify({'basarili': False, 'mesaj': 'Sipariş No gerekli'})

    returns = Return.query.filter(
        Return.siparis_no == siparis_no,
        Return.durum == 'HATALI',
    ).all()

    if not returns:
        return jsonify({'basarili': False, 'mesaj': 'Bu grupta HATALI iade bulunamadı'})

    duzeltilen = 0
    hatali_kalan = 0
    for ret in returns:
        errors = _revalidate_return(ret)
        if errors:
            ret.hata_sebebi = '; '.join(errors)
            hatali_kalan += 1
        else:
            ret.durum = 'BEKLEMEDE'
            ret.hata_sebebi = None
            duzeltilen += 1

    db.session.commit()
    log_audit('grup_revalidate', 'returns', None,
              yeni_deger={'siparis_no': siparis_no, 'duzeltilen': duzeltilen, 'hatali_kalan': hatali_kalan})

    if duzeltilen:
        mesaj = f'{duzeltilen} iade BEKLEMEDE oldu' + (f', {hatali_kalan} hata devam ediyor' if hatali_kalan else '')
    else:
        mesaj = f'Hatalar devam ediyor ({hatali_kalan} iade)'
    return jsonify({'basarili': True, 'duzeltilen': duzeltilen, 'hatali_kalan': hatali_kalan, 'mesaj': mesaj})


@iadeler_bp.route('/api/toplu-sil', methods=['POST'])
def api_toplu_sil():
    ids = (request.json or {}).get('ids', [])
    if not ids:
        return jsonify({'basarili': False, 'mesaj': 'ID listesi boş'})
    silinen = Return.query.filter(Return.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    log_audit('toplu_sil', 'returns', None, yeni_deger={'silinen': silinen})
    return jsonify({'basarili': True, 'mesaj': f'{silinen} iade silindi'})


@iadeler_bp.route('/api/toplu-guncelle', methods=['POST'])
def api_toplu_guncelle():
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'basarili': False, 'mesaj': 'ID listesi boş'})

    guncellenen = 0
    returns = Return.query.filter(Return.id.in_(ids)).all()
    for ret in returns:
        if 'toplama_id' in data and data['toplama_id']:
            ret.toplama_id = int(data['toplama_id'])
        guncellenen += 1

    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': f'{guncellenen} iade güncellendi'})


@iadeler_bp.route('/guncelle-hatali', methods=['POST'])
def guncelle_hatali_iadeler_route():
    """Tüm HATALI iadeleri yeniden doğrula."""
    hatali_iadeler = Return.query.filter_by(durum='HATALI').all()

    if not hatali_iadeler:
        return jsonify({
            'status': 'success',
            'basarili': True,
            'message': 'HATALI iade yok',
            'duzeltilen': 0,
            'hala_hatali': 0,
        })

    duzeltilen = 0
    hala_hatali = 0

    for ret in hatali_iadeler:
        errors = _revalidate_return(ret)
        if errors:
            ret.durum = 'HATALI'
            ret.hata_sebebi = '; '.join(errors)
            hala_hatali += 1
        else:
            ret.durum = 'BEKLEMEDE'
            ret.hata_sebebi = None
            duzeltilen += 1

    db.session.commit()
    log_audit('hatali_toplu_revalidate', 'returns', None,
              yeni_deger={'duzeltilen': duzeltilen, 'hatali_kalan': hala_hatali})

    return jsonify({
        'status': 'success',
        'basarili': True,
        'message': f"{duzeltilen} iade BEKLEMEDE'ye geçti. {hala_hatali} iade hâlâ HATALI",
        'duzeltilen': duzeltilen,
        'hala_hatali': hala_hatali,
    })


@iadeler_bp.route('/guncelle', methods=['POST'])
def tum_iadeleri_guncelle():
    """Seçili tarihteki tüm iadeleri yeniden doğrula ve güncelle."""
    data = request.json or {}
    tarih_str = str(data.get('tarih', '')).strip()
    if not tarih_str:
        return jsonify({'basarili': False, 'mesaj': 'Tarih gerekli'}), 400

    try:
        tarih = datetime.strptime(tarih_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'basarili': False, 'mesaj': 'Tarih formatı geçersiz'}), 400

    ertesi_gun = tarih + timedelta(days=1)
    iadeler = Return.query.filter(
        Return.tarih >= tarih,
        Return.tarih < ertesi_gun,
    ).all()

    if not iadeler:
        return jsonify({'basarili': False, 'mesaj': 'Bu tarihte iade bulunamadı'}), 404

    guncellenen = 0
    duzeltilen = 0
    hatali_kalan = 0

    for ret in iadeler:
        if ret.durum == 'TAMAMLANDI':
            guncellenen += 1
            continue

        onceki_durum = ret.durum
        errors = _revalidate_return(ret)

        if errors:
            ret.durum = 'HATALI'
            ret.hata_sebebi = '; '.join(errors)
            hatali_kalan += 1
        else:
            ret.durum = 'BEKLEMEDE'
            ret.hata_sebebi = None
            if onceki_durum != 'BEKLEMEDE':
                duzeltilen += 1

        guncellenen += 1

    db.session.commit()
    log_audit('tum_iadeler_guncellendi', 'returns', None, yeni_deger={
        'tarih': tarih_str,
        'guncellenen': guncellenen,
        'duzeltilen': duzeltilen,
        'hatali_kalan': hatali_kalan,
    })

    return jsonify({
        'basarili': True,
        'mesaj': f'{guncellenen} iade güncellendi. {duzeltilen} düzeltildi, {hatali_kalan} hatalı kaldı.',
        'guncellenen': guncellenen,
        'duzeltilen': duzeltilen,
        'hatali_kalan': hatali_kalan,
    })


@iadeler_bp.route('/listele', methods=['POST'])
def listele_uyumlu():
    """Eski/uyumlu endpoint: filtreye göre iadeleri döndür."""
    data = request.json or {}
    tarih_str = str(data.get('tarih', '')).strip()
    durum = str(data.get('durum', '')).strip()

    query = Return.query

    if tarih_str:
        try:
            tarih = datetime.strptime(tarih_str, '%Y-%m-%d')
            query = query.filter(
                Return.tarih >= tarih,
                Return.tarih < tarih + timedelta(days=1),
            )
        except ValueError:
            return jsonify({'status': 'error', 'message': 'Tarih formatı geçersiz'}), 400

    if durum:
        query = query.filter(Return.durum == durum)

    iadeler = query.order_by(Return.id.desc()).all()
    result = [{
        'id': r.id,
        'siparis_no': r.siparis_no,
        'tarih': r.tarih.strftime('%Y-%m-%d') if r.tarih else None,
        'urun_kodu': (r.urun.ana_kod if r.urun else r.urun_kodu_ham) or '',
        'beden': r.beden or '',
        'adet': r.adet,
        'sebep': r.sebebi or '',
        'toplama': r.toplama.ad if r.toplama else '',
        'durum': r.durum or 'BEKLEMEDE',
        'hata_sebebi': r.hata_sebebi or '',
    } for r in iadeler]

    return jsonify({'status': 'success', 'iadeler': result})


@iadeler_bp.route('/api/siparis-detay/<siparis_no>')
def api_siparis_detay(siparis_no):
    """Sipariş No'ya ait tüm iade satırlarını döndür."""
    returns = Return.query.filter_by(siparis_no=siparis_no).order_by(Return.id).all()
    if not returns:
        return jsonify({'basarili': False, 'mesaj': 'İade bulunamadı'}), 404

    first = returns[0]
    iade_dict = {
        'siparis_no': siparis_no,
        'tarih': first.tarih.strftime('%d.%m.%Y') if first.tarih else '-',
        'toplama': first.toplama.ad if first.toplama else '-',
        'durum': first.durum or 'BEKLEMEDE',
    }

    urunler = []
    for r in returns:
        urunler.append({
            'id': r.id,
            'urun_kodu': (r.urun.ana_kod if r.urun else r.urun_kodu_ham) or '-',
            'beden': r.beden or '-',
            'adet': r.adet,
            'sebep': r.sebebi or '-',
            'durum': r.durum,
            'hata_sebebi': r.hata_sebebi or '',
        })

    return jsonify({'basarili': True, 'iade': iade_dict, 'urunler': urunler})


@iadeler_bp.route('/api/gecmis')
def api_gecmis():
    uploads = ExcelUpload.query.filter_by(modul='iade').order_by(ExcelUpload.yukleme_tarihi.desc()).all()
    result = []
    for u in uploads:
        hatali_count = Return.query.filter_by(excel_yukleme_id=u.id, durum='HATALI').count()
        result.append({
            'id': u.id,
            'dosya_adi': u.dosya_adi,
            'yukleme_tarihi': u.yukleme_tarihi.strftime('%Y-%m-%d %H:%M'),
            'toplam_satir': u.toplam_satir,
            'basarili': u.basarili,
            'basarisiz': u.basarisiz,
            'hatali_count': hatali_count,
            'durum': u.durum,
        })
    return jsonify({'basarili': True, 'gecmis': result})


@iadeler_bp.route('/api/gecmis/<int:upload_id>', methods=['DELETE'])
def api_gecmis_sil(upload_id):
    upload = ExcelUpload.query.get_or_404(upload_id)
    if upload.modul != 'iade':
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz modül'}), 400
    Return.query.filter_by(excel_yukleme_id=upload_id).delete(synchronize_session=False)
    db.session.delete(upload)
    db.session.commit()
    log_audit('gecmis_sil', 'excel_uploads', upload_id)
    return jsonify({'basarili': True, 'mesaj': 'Yükleme ve iadeleri silindi'})


@iadeler_bp.route('/api/hata-excel/<int:upload_id>')
def api_hata_excel(upload_id):
    """Belirli yüklemeye ait HATALI iadeleri Excel olarak indir"""
    hatali = Return.query.filter_by(excel_yukleme_id=upload_id, durum='HATALI').all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hatalı İadeler'
    ws.append(['Satır No', 'Hata Sebebi', 'Sipariş No', 'Ürün Kodu', 'Beden', 'Adet', 'Sebep'])
    for i, r in enumerate(hatali, start=1):
        ws.append([
            i,
            r.hata_sebebi or '',
            r.siparis_no or '',
            (r.urun.ana_kod if r.urun else r.urun_kodu_ham) or '',
            r.beden or '',
            r.adet or 0,
            r.sebebi or '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'hatali_iadeler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@iadeler_bp.route('/api/hatali-excel-genel')
def api_hatali_excel_genel():
    """Tüm HATALI iadeleri Excel olarak indir"""
    hatali = Return.query.filter_by(durum='HATALI').order_by(Return.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hatalı İadeler'
    ws.append(['Satır No', 'Hata Sebebi', 'Sipariş No', 'Ürün Kodu', 'Beden', 'Adet', 'Sebep'])
    for i, r in enumerate(hatali, start=1):
        ws.append([
            i,
            r.hata_sebebi or '',
            r.siparis_no or '',
            (r.urun.ana_kod if r.urun else r.urun_kodu_ham) or '',
            r.beden or '',
            r.adet or 0,
            r.sebebi or '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'hatali_iadeler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@iadeler_bp.route('/api/excel-indir')
def api_excel_indir():
    """Filtrelenmiş iadeleri Excel olarak indir"""
    siparis_no = request.args.get('siparis_no', '').strip()
    urun_kodu = request.args.get('urun_kodu', '').strip()
    tarih_baslangic = request.args.get('tarih_baslangic', '').strip()
    tarih_bitis = request.args.get('tarih_bitis', '').strip()
    f_toplama_id = request.args.get('toplama_id', '').strip()
    f_durum = request.args.get('durum', '').strip()

    query = Return.query
    if siparis_no:
        query = query.filter(Return.siparis_no.ilike(f'%{siparis_no}%'))
    if urun_kodu:
        query = query.outerjoin(Product, Return.urun_id == Product.id).filter(
            or_(
                Product.ana_kod.ilike(f'%{urun_kodu}%'),
                Return.urun_kodu_ham.ilike(f'%{urun_kodu}%'),
            )
        )
    if tarih_baslangic:
        try:
            query = query.filter(Return.tarih >= datetime.strptime(tarih_baslangic, '%Y-%m-%d'))
        except ValueError:
            pass
    if tarih_bitis:
        try:
            query = query.filter(Return.tarih < datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass
    if f_toplama_id:
        query = query.filter(Return.toplama_id == int(f_toplama_id))
    if f_durum:
        query = query.filter(Return.durum == f_durum)

    returns = query.order_by(Return.tarih.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'İadeler'
    ws.append(['Sipariş No', 'Tarih', 'Ürün Kodu', 'Beden', 'Adet', 'Sebep', 'Toplama', 'Durum'])
    for r in returns:
        ws.append([
            r.siparis_no or '',
            r.tarih.strftime('%d.%m.%Y') if r.tarih else '',
            (r.urun.ana_kod if r.urun else r.urun_kodu_ham) or '',
            r.beden or '',
            r.adet or 0,
            r.sebebi or '',
            r.toplama.ad if r.toplama else '',
            r.durum or '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'iadeler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@iadeler_bp.route('/api/analizler')
def api_analizler():
    from sqlalchemy import func

    # En fazla iade edilen ürünler
    urun_iadeleri = db.session.query(
        Product.ana_kod,
        Product.marka,
        func.coalesce(func.sum(Return.adet), 0).label('adet'),
    ).outerjoin(Return, Return.urun_id == Product.id).group_by(Product.id).order_by(func.sum(Return.adet).desc()).limit(10).all()

    # İade nedenleri
    nedenler = db.session.query(
        Return.sebebi,
        func.coalesce(func.sum(Return.adet), 0).label('adet'),
    ).group_by(Return.sebebi).all()

    # Toplama bazlı iadeler
    toplama_iadeleri = db.session.query(
        Toplama.ad.label('toplama'),
        func.coalesce(func.sum(Return.adet), 0).label('adet'),
    ).outerjoin(Return, Return.toplama_id == Toplama.id).group_by(Toplama.id).all()

    return jsonify({
        'basarili': True,
        'urun_iadeleri': [{'kod': r.ana_kod, 'marka': r.marka, 'adet': int(r.adet)} for r in urun_iadeleri],
        'iade_nedenleri': [{'sebep': r.sebebi or 'Belirtilmedi', 'adet': int(r.adet)} for r in nedenler],
        'toplama_iadeleri': [{'toplama': r.toplama, 'adet': int(r.adet)} for r in toplama_iadeleri],
    })

