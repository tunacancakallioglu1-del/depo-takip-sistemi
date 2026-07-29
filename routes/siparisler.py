# -*- coding: utf-8 -*-
"""Sipariş yönetimi rotaları — E-Ticaret Admin Panel"""

from io import BytesIO
from datetime import datetime, timedelta

from flask import Blueprint, render_template, request, jsonify, send_file
from openpyxl import Workbook
from sqlalchemy import or_

from database import db, Order, Product, Toplama, ExcelUpload
from routes.excel_handler import match_product_by_code, determine_toplama, parse_row_date
from utils.excel_utils import build_template, load_excel_rows, calculate_file_hash
from utils.audit_utils import log_audit

siparisler_bp = Blueprint('siparisler', __name__, url_prefix='/siparisler')

SIPARIS_HEADERS = ['Siparis No', 'Tarih', 'Urun Kodu', 'Beden', 'Adet', 'Toplama']


def _check_order(siparis_no, tarih, urun_id, urun_kodu_ham, beden, adet, toplama_id, product=None):
    """Returns list of error strings for an order."""
    errors = []
    if not siparis_no or str(siparis_no).strip() == '':
        errors.append('Sipariş No boş')
    if not tarih:
        errors.append('Tarih boş')
    if not urun_id:
        if urun_kodu_ham:
            errors.append(f'Ürün tanımsız: {urun_kodu_ham}')
        else:
            errors.append('Ürün Kodu boş')
    if product and product.beden_ayrimi and not beden:
        errors.append('Beden boş/tanımsız')
    try:
        adet_int = int(adet or 0)
    except (ValueError, TypeError):
        adet_int = 0
    if adet_int <= 0:
        errors.append('Adet ≤ 0')
    if not toplama_id:
        errors.append('Toplama seçilmemiş')
    return errors


@siparisler_bp.route('/')
def index():
    return render_template('siparisler.html')


@siparisler_bp.route('/api/meta')
def api_meta():
    """Dropdown verileri: toplamalar"""
    toplamalar = Toplama.query.order_by(Toplama.ad).all()
    return jsonify({
        'basarili': True,
        'toplamalar': [{'id': t.id, 'ad': t.ad} for t in toplamalar],
    })


@siparisler_bp.route('/api/template')
def download_template():
    stream = build_template(SIPARIS_HEADERS)
    return send_file(
        stream, as_attachment=True,
        download_name='siparis_sablonu.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@siparisler_bp.route('/api/excel-yukle', methods=['POST'])
def excel_yukle():
    file = request.files.get('file')
    if not file:
        return jsonify({'basarili': False, 'mesaj': 'Dosya zorunludur!'}), 400

    filename = file.filename or ''
    if not filename.lower().endswith(('.xlsx', '.xlsm', '.xltx', '.xltm')):
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz dosya formatı! (.xlsx gerekli)'}), 400

    file_hash = calculate_file_hash(file)
    if ExcelUpload.query.filter_by(dosya_hash=file_hash).first():
        return jsonify({'basarili': False, 'mesaj': 'Bu Excel dosyası daha önce yüklenmiş!'}), 409

    try:
        headers, rows = load_excel_rows(file)
    except Exception:
        return jsonify({'basarili': False, 'mesaj': 'Excel okunamadı. Dosya formatını kontrol edin.'}), 400

    # Zorunlu başlıkları kontrol et; fazladan sütunlar (örn. Personel) yoksayılır
    actual_core = [str(h).strip() for h in headers[:len(SIPARIS_HEADERS)]]
    if actual_core != SIPARIS_HEADERS:
        return jsonify({
            'basarili': False,
            'mesaj': 'Excel başlıkları/sırası hatalı!',
            'beklenen': SIPARIS_HEADERS,
            'bulunan': headers,
        }), 400

    upload = ExcelUpload(
        modul='siparis', dosya_adi=filename,
        dosya_hash=file_hash, toplam_satir=len(rows),
    )
    db.session.add(upload)
    db.session.flush()
    log_audit('excel_yukleme_basladi', 'excel_uploads', upload.id,
              yeni_deger={'modul': 'siparis', 'dosya': filename})

    basarili = 0
    hatali = 0
    row_errors = []
    unique_orders = set()

    for row_no, row in enumerate(rows, start=2):
        siparis_no = str(row.get('Siparis No', '') or '').strip()
        urun_kodu_raw = str(row.get('Urun Kodu', '') or '').strip()
        beden = str(row.get('Beden', '') or '').strip() or None
        toplama_raw = str(row.get('Toplama', '') or '').strip() or None
        # Personel sütunu siparişlerde kullanılmıyor — yoksay

        try:
            adet = int(float(row.get('Adet') or 0))
        except (ValueError, TypeError):
            adet = 0

        tarih = parse_row_date(row.get('Tarih'))

        errors = []
        product = None
        urun_id = None
        toplama_id = None

        if not siparis_no:
            errors.append('Sipariş No boş')

        if not urun_kodu_raw:
            errors.append('Ürün Kodu boş')
        else:
            product = match_product_by_code(urun_kodu_raw)
            if not product:
                errors.append(f'Ürün tanımsız: {urun_kodu_raw}')
            else:
                urun_id = product.id
                tid, beden_error = determine_toplama(product, beden)
                if beden_error:
                    errors.append(beden_error)
                else:
                    toplama_id = tid

        # Toplama sütunundan da dene (ürün bulunamadığında veya beden hatası olmadığında)
        if not toplama_id and toplama_raw:
            t = Toplama.query.filter_by(ad=toplama_raw).first()
            if t:
                toplama_id = t.id

        if not toplama_id and not any('Toplama' in e or 'beden' in e.lower() or 'Beden' in e for e in errors):
            errors.append('Toplama seçilmemiş')

        if adet <= 0:
            errors.append('Adet ≤ 0')

        if not tarih:
            errors.append('Tarih boş')

        if errors:
            durum = 'HATALI'
            hata_sebebi = '; '.join(errors)
            hatali += 1
            row_errors.append({
                'satir': row_no,
                'siparis_no': siparis_no,
                'urun_kodu': urun_kodu_raw,
                'beden': beden or '',
                'adet': adet,
                'toplama': toplama_raw or '',
                'hata': hata_sebebi,
            })
        else:
            durum = 'BEKLEMEDE'
            hata_sebebi = None
            basarili += 1

        # Aynı sipariş no + urun kodu ile HATALI kayıt varsa sil (yeniden yükleme)
        if siparis_no:
            existing_q = Order.query.filter(
                Order.siparis_no == siparis_no,
                Order.durum == 'HATALI',
            )
            if urun_kodu_raw:
                existing_q = existing_q.filter(
                    or_(Order.urun_kodu_ham == urun_kodu_raw,
                        Order.urun_id == urun_id) if urun_id else Order.urun_kodu_ham == urun_kodu_raw
                )
            existing = existing_q.first()
            if existing:
                db.session.delete(existing)
                db.session.flush()

        order = Order(
            siparis_no=siparis_no or '?',
            tarih=tarih or datetime.utcnow(),
            urun_id=urun_id,
            urun_kodu_ham=urun_kodu_raw or None,
            beden=beden,
            adet=adet,
            toplama_id=toplama_id,
            personel_id=None,
            durum=durum,
            hata_sebebi=hata_sebebi,
            excel_yukleme_id=upload.id,
        )
        db.session.add(order)
        if siparis_no:
            unique_orders.add(siparis_no)

    upload.basarili = basarili
    upload.basarisiz = hatali
    log_audit('excel_yukleme_tamamlandi', 'orders', upload.id,
              yeni_deger={'toplam_satir': len(rows), 'basarili': basarili, 'hatali': hatali})
    db.session.commit()

    return jsonify({
        'basarili': True,
        'mesaj': 'Sipariş yükleme tamamlandı',
        'upload_id': upload.id,
        'ozet': {
            'toplam_satir': len(rows),
            'tekil_siparis': len(unique_orders),
            'islenen_satir': basarili,
            'hatali_satir': hatali,
        },
        'hatalar': row_errors,
    })


@siparisler_bp.route('/api/list')
def api_list():
    page = max(int(request.args.get('page', 1)), 1)
    per_page = min(max(int(request.args.get('per_page', 20)), 1), 100)

    siparis_no = request.args.get('siparis_no', '').strip()
    urun_kodu = request.args.get('urun_kodu', '').strip()
    tarih_baslangic = request.args.get('tarih_baslangic', '').strip()
    tarih_bitis = request.args.get('tarih_bitis', '').strip()
    f_toplama_id = request.args.get('toplama_id', '').strip()
    f_durum = request.args.get('durum', '').strip()

    query = Order.query

    if siparis_no:
        query = query.filter(Order.siparis_no.ilike(f'%{siparis_no}%'))

    if urun_kodu:
        query = query.outerjoin(Product, Order.urun_id == Product.id).filter(
            or_(
                Product.ana_kod.ilike(f'%{urun_kodu}%'),
                Order.urun_kodu_ham.ilike(f'%{urun_kodu}%'),
            )
        )

    if tarih_baslangic:
        try:
            query = query.filter(Order.tarih >= datetime.strptime(tarih_baslangic, '%Y-%m-%d'))
        except ValueError:
            pass

    if tarih_bitis:
        try:
            dt_end = datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Order.tarih < dt_end)
        except ValueError:
            pass

    if f_toplama_id:
        query = query.filter(Order.toplama_id == int(f_toplama_id))

    if f_durum:
        query = query.filter(Order.durum == f_durum)

    pagination = query.order_by(Order.id.desc()).paginate(page=page, per_page=per_page, error_out=False)

    rows = []
    for item in pagination.items:
        rows.append({
            'id': item.id,
            'siparis_no': item.siparis_no,
            'tarih': item.tarih.strftime('%Y-%m-%d') if item.tarih else None,
            'urun_kodu': (item.urun.ana_kod if item.urun else item.urun_kodu_ham) or '',
            'beden': item.beden or '',
            'adet': item.adet,
            'toplama': item.toplama.ad if item.toplama else '',
            'toplama_id': item.toplama_id,
            'durum': item.durum or 'BEKLEMEDE',
            'hata_sebebi': item.hata_sebebi or '',
            'senkronize_edildi': item.senkronize_edildi,
        })

    return jsonify({'basarili': True, 'kayitlar': rows, 'toplam': pagination.total})


@siparisler_bp.route('/api/<int:order_id>', methods=['GET'])
def api_get(order_id):
    order = Order.query.get_or_404(order_id)
    return jsonify({
        'basarili': True,
        'siparis': {
            'id': order.id,
            'siparis_no': order.siparis_no,
            'tarih': order.tarih.strftime('%Y-%m-%d') if order.tarih else None,
            'urun_kodu': (order.urun.ana_kod if order.urun else order.urun_kodu_ham) or '',
            'beden': order.beden or '',
            'adet': order.adet,
            'toplama_id': order.toplama_id,
            'durum': order.durum or 'BEKLEMEDE',
            'hata_sebebi': order.hata_sebebi or '',
        },
    })


@siparisler_bp.route('/api/<int:order_id>', methods=['PUT'])
def api_guncelle(order_id):
    order = Order.query.get_or_404(order_id)
    data = request.json or {}
    eski_durum = order.durum
    eski = {'siparis_no': order.siparis_no, 'beden': order.beden, 'adet': order.adet, 'durum': order.durum}

    if 'beden' in data:
        order.beden = str(data['beden'] or '').strip() or None

    if 'adet' in data:
        try:
            order.adet = int(float(data['adet'] or 0))
        except (ValueError, TypeError):
            order.adet = 0

    if 'toplama_id' in data and data['toplama_id']:
        order.toplama_id = int(data['toplama_id'])

    if 'tarih' in data and data['tarih']:
        try:
            order.tarih = datetime.strptime(str(data['tarih']), '%Y-%m-%d')
        except ValueError:
            pass

    if 'urun_kodu' in data and data['urun_kodu']:
        urun_kodu = str(data['urun_kodu']).strip()
        order.urun_kodu_ham = urun_kodu
        product = match_product_by_code(urun_kodu)
        if product:
            order.urun_id = product.id
            if 'toplama_id' not in data or not data['toplama_id']:
                tid, _ = determine_toplama(product, order.beden)
                if tid:
                    order.toplama_id = tid
        else:
            order.urun_id = None

    # Durum otomatik belirleme (TAMAMLANDI ise değiştirme)
    if eski_durum != 'TAMAMLANDI':
        errors = _check_order(
            siparis_no=order.siparis_no,
            tarih=order.tarih,
            urun_id=order.urun_id,
            urun_kodu_ham=order.urun_kodu_ham,
            beden=order.beden,
            adet=order.adet,
            toplama_id=order.toplama_id,
            product=order.urun,
        )
        if errors:
            order.durum = 'HATALI'
            order.hata_sebebi = '; '.join(errors)
        else:
            order.durum = 'BEKLEMEDE'
            order.hata_sebebi = None

    log_audit('update', 'orders', order_id, eski_deger=eski,
              yeni_deger={'beden': order.beden, 'adet': order.adet, 'durum': order.durum})
    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': 'Sipariş güncellendi', 'yeni_durum': order.durum})


@siparisler_bp.route('/api/<int:order_id>', methods=['DELETE'])
def api_sil(order_id):
    order = Order.query.get_or_404(order_id)
    log_audit('delete', 'orders', order_id, eski_deger={'siparis_no': order.siparis_no})
    db.session.delete(order)
    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': 'Sipariş silindi'})


@siparisler_bp.route('/api/toplu-sil', methods=['POST'])
def api_toplu_sil():
    ids = (request.json or {}).get('ids', [])
    if not ids:
        return jsonify({'basarili': False, 'mesaj': 'ID listesi boş'})
    silinen = Order.query.filter(Order.id.in_(ids)).delete(synchronize_session=False)
    db.session.commit()
    log_audit('toplu_sil', 'orders', None, yeni_deger={'silinen': silinen})
    return jsonify({'basarili': True, 'mesaj': f'{silinen} sipariş silindi'})


@siparisler_bp.route('/api/toplu-guncelle', methods=['POST'])
def api_toplu_guncelle():
    data = request.json or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'basarili': False, 'mesaj': 'ID listesi boş'})

    guncellenen = 0
    orders = Order.query.filter(Order.id.in_(ids)).all()
    for order in orders:
        if 'toplama_id' in data and data['toplama_id']:
            order.toplama_id = int(data['toplama_id'])
        guncellenen += 1

    db.session.commit()
    return jsonify({'basarili': True, 'mesaj': f'{guncellenen} sipariş güncellendi'})


@siparisler_bp.route('/api/gecmis')
def api_gecmis():
    uploads = ExcelUpload.query.filter_by(modul='siparis').order_by(
        ExcelUpload.yukleme_tarihi.desc()
    ).all()
    result = []
    for u in uploads:
        hatali_count = Order.query.filter_by(excel_yukleme_id=u.id, durum='HATALI').count()
        result.append({
            'id': u.id,
            'dosya_adi': u.dosya_adi,
            'yukleme_tarihi': u.yukleme_tarihi.strftime('%Y-%m-%d %H:%M'),
            'toplam_satir': u.toplam_satir,
            'basarili': u.basarili,
            'basarisiz': u.basarisiz,
            'hatali_count': hatali_count,
            'durum': u.durum,
        })
    return jsonify({'basarili': True, 'gecmis': result})


@siparisler_bp.route('/api/gecmis/<int:upload_id>', methods=['DELETE'])
def api_gecmis_sil(upload_id):
    upload = ExcelUpload.query.get_or_404(upload_id)
    if upload.modul != 'siparis':
        return jsonify({'basarili': False, 'mesaj': 'Geçersiz modül'}), 400
    Order.query.filter_by(excel_yukleme_id=upload_id).delete(synchronize_session=False)
    db.session.delete(upload)
    db.session.commit()
    log_audit('gecmis_sil', 'excel_uploads', upload_id)
    return jsonify({'basarili': True, 'mesaj': 'Yükleme ve siparişleri silindi'})


@siparisler_bp.route('/api/hata-excel/<int:upload_id>')
def api_hata_excel(upload_id):
    """Belirli yüklemeye ait HATALI siparişleri Excel olarak indir"""
    hatali = Order.query.filter_by(excel_yukleme_id=upload_id, durum='HATALI').all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hatalı Siparişler'
    ws.append(['Satır No', 'Hata Sebebi', 'Sipariş No', 'Ürün Kodu', 'Beden', 'Adet', 'Toplama'])
    for i, o in enumerate(hatali, start=1):
        ws.append([
            i,
            o.hata_sebebi or '',
            o.siparis_no or '',
            (o.urun.ana_kod if o.urun else o.urun_kodu_ham) or '',
            o.beden or '',
            o.adet or 0,
            o.toplama.ad if o.toplama else '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'hatali_siparisler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@siparisler_bp.route('/api/hatali-excel-genel')
def api_hatali_excel_genel():
    """Tüm HATALI siparişleri Excel olarak indir"""
    hatali = Order.query.filter_by(durum='HATALI').order_by(Order.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Hatalı Siparişler'
    ws.append(['Satır No', 'Hata Sebebi', 'Sipariş No', 'Ürün Kodu', 'Beden', 'Adet', 'Toplama'])
    for i, o in enumerate(hatali, start=1):
        ws.append([
            i,
            o.hata_sebebi or '',
            o.siparis_no or '',
            (o.urun.ana_kod if o.urun else o.urun_kodu_ham) or '',
            o.beden or '',
            o.adet or 0,
            o.toplama.ad if o.toplama else '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'hatali_siparisler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@siparisler_bp.route('/api/excel-indir')
def api_excel_indir():
    """Filtrelenmiş siparişleri Excel olarak indir"""
    siparis_no = request.args.get('siparis_no', '').strip()
    urun_kodu = request.args.get('urun_kodu', '').strip()
    tarih_baslangic = request.args.get('tarih_baslangic', '').strip()
    tarih_bitis = request.args.get('tarih_bitis', '').strip()
    f_toplama_id = request.args.get('toplama_id', '').strip()
    f_durum = request.args.get('durum', '').strip()

    query = Order.query
    if siparis_no:
        query = query.filter(Order.siparis_no.ilike(f'%{siparis_no}%'))
    if urun_kodu:
        query = query.outerjoin(Product, Order.urun_id == Product.id).filter(
            or_(
                Product.ana_kod.ilike(f'%{urun_kodu}%'),
                Order.urun_kodu_ham.ilike(f'%{urun_kodu}%'),
            )
        )
    if tarih_baslangic:
        try:
            query = query.filter(Order.tarih >= datetime.strptime(tarih_baslangic, '%Y-%m-%d'))
        except ValueError:
            pass
    if tarih_bitis:
        try:
            query = query.filter(Order.tarih < datetime.strptime(tarih_bitis, '%Y-%m-%d') + timedelta(days=1))
        except ValueError:
            pass
    if f_toplama_id:
        query = query.filter(Order.toplama_id == int(f_toplama_id))
    if f_durum:
        query = query.filter(Order.durum == f_durum)

    orders = query.order_by(Order.tarih.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = 'Siparişler'
    ws.append(['Sipariş No', 'Tarih', 'Ürün Kodu', 'Beden', 'Adet', 'Toplama', 'Durum'])
    for o in orders:
        ws.append([
            o.siparis_no or '',
            o.tarih.strftime('%d.%m.%Y') if o.tarih else '',
            (o.urun.ana_kod if o.urun else o.urun_kodu_ham) or '',
            o.beden or '',
            o.adet or 0,
            o.toplama.ad if o.toplama else '',
            o.durum or '',
        ])

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    tarih_str = datetime.now().strftime('%d_%m_%Y')
    return send_file(
        out, as_attachment=True,
        download_name=f'siparisler_{tarih_str}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

