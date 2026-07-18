# -*- coding: utf-8 -*-
"""
Raporlar Rotaları
"""

from flask import Blueprint, render_template, request, jsonify, send_file
from database import db, Kayit, Personel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import Config
import os
from datetime import datetime

raporlar_bp = Blueprint('raporlar', __name__, url_prefix='/raporlar')


@raporlar_bp.route('/')
def index():
    """Raporlar sayfası"""
    
    return render_template('raporlar.html')


@raporlar_bp.route('/api/excel-olustur', methods=['POST'])
def api_excel_olustur():
    """AJAX ile Excel raporu oluştur"""
    
    try:
        veri = request.json
        
        # Tarih aralığını kontrol et
        if not veri.get('tarih_baslangic') or not veri.get('tarih_bitis'):
            return jsonify({
                'basarili': False,
                'mesaj': 'Başlangıç ve bitiş tarihleri seçmelisiniz!'
            }), 400
        
        tarih_baslangic = veri['tarih_baslangic']
        tarih_bitis = veri['tarih_bitis']
        
        # Tarihları DD.MM.YYYY formatına dönüştür
        parts_bas = tarih_baslangic.split('-')
        parts_bit = tarih_bitis.split('-')
        tarih_bas_fmt = f"{parts_bas[2]}.{parts_bas[1]}.{parts_bas[0]}"
        tarih_bit_fmt = f"{parts_bit[2]}.{parts_bit[1]}.{parts_bit[0]}"
        
        # Kayıtları sorgula
        kayitlar = Kayit.query.filter(
            Kayit.tarih >= tarih_bas_fmt,
            Kayit.tarih <= tarih_bit_fmt
        ).order_by(Kayit.tarih.desc()).all()
        
        if not kayitlar:
            return jsonify({
                'basarili': False,
                'mesaj': 'Seçilen tarih aralığında kayıt bulunamadı!'
            }), 400
        
        # Excel oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = 'Kayıtlar'
        
        # Başlık satırı
        basliklar = ['Tarih', 'Personel', 'Toplama', 'Trendyol Sipariş', 
                     'Trendyol Fatura', 'Diğer Pazar', 'Not']
        ws.append(basliklar)
        
        # Başlık stili
        baslik_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        baslik_font = Font(bold=True, color='FFFFFF')
        
        for cell in ws[1]:
            cell.fill = baslik_fill
            cell.font = baslik_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Kayıtları ekle
        for kayit in kayitlar:
            ws.append([
                kayit.tarih,
                kayit.personel.ad,
                kayit.toplama.ad,
                kayit.trendyol_siparis,
                kayit.trendyol_fatura,
                kayit.diger_pazar,
                kayit.not_alan
            ])
        
        # Sütun genişlikleri
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 20
        
        # Personel bazında toplamalar
        personeller = Personel.query.all()
        
        # Boş satır
        ws.append([])
        
        # Özet başlığı
        ozet_satir = ws.max_row + 1
        ws[f'A{ozet_satir}'] = 'PERSONEL BAZINDA ÖZET'
        ws[f'A{ozet_satir}'].font = Font(bold=True, size=12)
        
        # Personel özetleri
        for personel in personeller:
            kayitlar_personel = [k for k in kayitlar if k.personel_id == personel.id]
            
            if kayitlar_personel:
                ws.append([])
                ws.append([personel.ad])
                
                # Personel başlığı
                personel_baslik_satir = ws.max_row
                ws[f'A{personel_baslik_satir}'].font = Font(bold=True, color='FFFFFF')
                ws[f'A{personel_baslik_satir}'].fill = PatternFill(start_color='70AD47', 
                                                                   end_color='70AD47', 
                                                                   fill_type='solid')
                
                # Özet bilgileri
                toplam_siparis = sum(k.trendyol_siparis for k in kayitlar_personel)
                toplam_fatura = sum(k.trendyol_fatura for k in kayitlar_personel)
                toplam_diger = sum(k.diger_pazar for k in kayitlar_personel)
                
                ws.append(['Toplam Trendyol Sipariş', toplam_siparis])
                ws.append(['Toplam Trendyol Fatura', toplam_fatura])
                ws.append(['Toplam Diğer Pazar', toplam_diger])
        
        # Dosyayı kaydet
        dosya_adi = f"rapor_{datetime.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        dosya_yolu = os.path.join(Config.EXPORT_FOLDER, dosya_adi)
        
        wb.save(dosya_yolu)
        
        return jsonify({
            'basarili': True,
            'mesaj': 'Excel raporu başarıyla oluşturuldu!',
            'dosya': dosya_adi
        })
    
    except Exception as e:
        return jsonify({
            'basarili': False,
            'mesaj': f'Hata: {str(e)}'
        }), 500


@raporlar_bp.route('/download/<dosya_adi>')
def download(dosya_adi):
    """Excel dosyasını indir"""
    
    try:
        dosya_yolu = os.path.join(Config.EXPORT_FOLDER, dosya_adi)
        
        if not os.path.exists(dosya_yolu):
            return jsonify({'basarili': False, 'mesaj': 'Dosya bulunamadı!'}), 404
        
        return send_file(dosya_yolu, as_attachment=True)
    
    except Exception as e:
        return jsonify({'basarili': False, 'mesaj': f'Hata: {str(e)}'}), 500
