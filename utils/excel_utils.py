# -*- coding: utf-8 -*-
"""Excel işleme yardımcıları"""

import hashlib
from io import BytesIO
from openpyxl import load_workbook, Workbook


def calculate_file_hash(file_storage):
    data = file_storage.read()
    file_storage.seek(0)
    return hashlib.sha256(data).hexdigest()


def load_excel_rows(file_storage):
    workbook = load_workbook(filename=BytesIO(file_storage.read()), data_only=True)
    file_storage.seek(0)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    headers = [str(h).strip() if h is not None else '' for h in rows[0]] if rows else []
    data = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        data.append({headers[i]: row[i] for i in range(len(headers))})
    return headers, data


def build_template(headers):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Template'
    ws.append(headers)
    for idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=idx).value = header
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out
